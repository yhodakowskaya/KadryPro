from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from django.db.models import Q
from django.core.management import call_command
from django.core.mail import send_mail
from django.conf import settings
from io import StringIO
from datetime import date
from .models import User, Department, Position, CustomRole, PasswordResetToken, Company, Region, Contract
from .serializers import (
    CustomTokenObtainPairSerializer, UserListSerializer, UserDetailSerializer,
    UserCreateSerializer, UserUpdateSerializer, SetPasswordSerializer,
    DepartmentSerializer, DepartmentTreeSerializer,
    PositionSerializer, CustomRoleSerializer,
    CompanySerializer, RegionSerializer, ContractSerializer,
)
from .permissions import IsHROrAdmin, IsAdminOnly


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(UserDetailSerializer(request.user).data)

    def put(self, request):
        serializer = UserUpdateSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserDetailSerializer(request.user).data)


class UserListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = User.objects.select_related('department', 'manager', 'custom_role').all()
        role = self.request.query_params.get('role')
        dept = self.request.query_params.get('department')
        search = self.request.query_params.get('search')
        is_active = self.request.query_params.get('is_active')

        if role:
            qs = qs.filter(role=role)
        if dept:
            qs = qs.filter(department_id=dept)
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search) | Q(last_name__icontains=search) |
                Q(email__icontains=search) | Q(position__icontains=search)
            )
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        return qs

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserCreateSerializer
        return UserListSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsHROrAdmin()]
        return [IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        serializer = UserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserDetailSerializer(user).data, status=status.HTTP_201_CREATED)


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.select_related('department', 'manager', 'substitute_manager', 'custom_role').all()

    def get_serializer_class(self):
        if self.request.method in ('PUT', 'PATCH'):
            return UserUpdateSerializer
        return UserDetailSerializer

    def get_permissions(self):
        if self.request.method in ('PUT', 'PATCH', 'DELETE'):
            return [IsHROrAdmin()]
        return [IsAuthenticated()]

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = UserUpdateSerializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserDetailSerializer(instance).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance == request.user:
            return Response({'detail': 'Nie możesz usunąć własnego konta.'}, status=400)
        instance.is_active = False
        if not instance.termination_date:
            instance.termination_date = date.today()
        instance.save(update_fields=['is_active', 'termination_date'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class UserRestoreView(APIView):
    permission_classes = [IsAdminOnly]

    def post(self, request, pk):
        user = generics.get_object_or_404(User, pk=pk)
        user.is_active = True
        user.termination_date = None
        user.save(update_fields=['is_active', 'termination_date'])
        return Response(UserDetailSerializer(user).data)


class SetPasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        target = generics.get_object_or_404(User, pk=pk)
        if request.user.pk != target.pk and request.user.role != User.ROLE_ADMIN:
            return Response({'detail': 'Brak uprawnień do zmiany hasła innego użytkownika.'}, status=403)
        serializer = SetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        target.set_password(serializer.validated_data['password'])
        target.save()
        return Response({'detail': 'Hasło zostało zmienione.'})


# ── Departments ──────────────────────────────────────────────────────────────

class DepartmentListCreateView(generics.ListCreateAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsHROrAdmin()]
        return [IsAuthenticated()]


class DepartmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsHROrAdmin]


class DepartmentTreeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        roots = Department.objects.filter(parent=None)
        return Response(DepartmentTreeSerializer(roots, many=True).data)


# ── Positions ─────────────────────────────────────────────────────────────────

class PositionListCreateView(generics.ListCreateAPIView):
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = Position.objects.all()
        if self.request.query_params.get('active_only') == 'true':
            qs = qs.filter(is_active=True)
        return qs

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsHROrAdmin()]
        return [IsAuthenticated()]


class PositionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    permission_classes = [IsHROrAdmin]


# ── Custom Roles ──────────────────────────────────────────────────────────────

class CustomRoleListCreateView(generics.ListCreateAPIView):
    queryset = CustomRole.objects.all()
    serializer_class = CustomRoleSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdminOnly()]
        return [IsHROrAdmin()]


class CustomRoleDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CustomRole.objects.all()
    serializer_class = CustomRoleSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsHROrAdmin()]
        return [IsAdminOnly()]


# ── Password Reset ────────────────────────────────────────────────────────────

class ChangeOwnPasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        old = request.data.get('old_password', '')
        new = request.data.get('new_password', '')
        if not request.user.check_password(old):
            return Response({'detail': 'Aktualne hasło jest nieprawidłowe.'}, status=400)
        if len(new) < 6:
            return Response({'detail': 'Nowe hasło musi mieć co najmniej 6 znaków.'}, status=400)
        request.user.set_password(new)
        request.user.save()
        return Response({'detail': 'Hasło zostało zmienione.'})


class PasswordResetRequestView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email', '').strip()
        try:
            user = User.objects.get(email__iexact=email, is_active=True)
        except User.DoesNotExist:
            return Response({'detail': 'Jeśli email jest w systemie, wyślemy link.'})

        PasswordResetToken.objects.filter(user=user, used=False).update(used=True)
        token = PasswordResetToken.objects.create(user=user)
        reset_url = f"{settings.APP_URL}/reset-hasla/{token.token}"
        try:
            send_mail(
                subject='Reset hasła — KadryPro',
                message=(
                    f'Cześć {user.first_name},\n\n'
                    f'Otrzymaliśmy prośbę o reset hasła do systemu KadryPro.\n\n'
                    f'Kliknij poniższy link, aby ustawić nowe hasło:\n{reset_url}\n\n'
                    f'Link jest ważny przez 1 godzinę.\n\n'
                    f'Jeśli to nie Ty, zignoruj tę wiadomość.\n\nZespół KadryPro'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
        except Exception:
            pass
        return Response({'detail': 'Jeśli email jest w systemie, wyślemy link.'})


class PasswordResetConfirmView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        token_str = request.data.get('token', '')
        password = request.data.get('password', '')
        if len(password) < 6:
            return Response({'detail': 'Hasło musi mieć co najmniej 6 znaków.'}, status=400)
        try:
            token = PasswordResetToken.objects.select_related('user').get(token=token_str)
        except PasswordResetToken.DoesNotExist:
            return Response({'detail': 'Nieprawidłowy lub wygasły link.'}, status=400)
        if not token.is_valid():
            return Response({'detail': 'Link wygasł lub był już użyty. Poproś o nowy.'}, status=400)
        token.user.set_password(password)
        token.user.save()
        token.used = True
        token.save()
        return Response({'detail': 'Hasło zostało zmienione. Możesz się zalogować.'})


# ── Notifications ─────────────────────────────────────────────────────────────

class TriggerNotificationsView(APIView):
    permission_classes = [IsHROrAdmin]

    def post(self, request):
        out = StringIO()
        call_command('check_notifications', stdout=out)
        return Response({'detail': out.getvalue() or 'Sprawdzono powiadomienia.'})


# ── Org Chart ─────────────────────────────────────────────────────────────────

class OrgChartView(APIView):
    permission_classes = [IsAuthenticated]

    ROLE_DISPLAY = {
        'admin': 'Administrator',
        'hr': 'Kadry/HR',
        'manager': 'Kierownik',
        'employee': 'Pracownik',
    }

    def get(self, request):
        from hr.models import VacationRequest
        from datetime import date

        today = date.today()
        active_vacations = {
            v['employee_id']: v
            for v in VacationRequest.objects.filter(
                status=VacationRequest.STATUS_APPROVED,
                start_date__lte=today,
                end_date__gte=today,
            ).values('employee_id', 'start_date', 'end_date', 'request_type',
                     'vacation_type__name', 'vacation_type__color')
        }

        users_qs = User.objects.filter(is_active=True).select_related(
            'department', 'manager', 'custom_role'
        ).values(
            'id', 'first_name', 'last_name', 'position', 'role',
            'department__name', 'manager_id', 'custom_role__name', 'custom_role__color'
        )

        result = []
        for u in users_qs:
            row = dict(u)
            row['role_display'] = self.ROLE_DISPLAY.get(u['role'], u['role'])
            vac = active_vacations.get(u['id'])
            if vac:
                row['on_vacation'] = True
                row['vacation_start'] = vac['start_date'].strftime('%d.%m.%Y')
                row['vacation_end'] = vac['end_date'].strftime('%d.%m.%Y')
                row['vacation_type_name'] = vac.get('vacation_type__name') or (
                    'Praca zdalna' if vac['request_type'] == 'remote' else 'Urlop'
                )
                row['vacation_type_color'] = vac.get('vacation_type__color') or (
                    'purple' if vac['request_type'] == 'remote' else 'amber'
                )
            else:
                row['on_vacation'] = False
                row['vacation_type_name'] = None
                row['vacation_type_color'] = None
            result.append(row)

        return Response(result)


# ── Excel Import ──────────────────────────────────────────────────────────────

class DepartmentImportView(APIView):
    permission_classes = [IsHROrAdmin]

    def post(self, request):
        import openpyxl
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Nie przesłano pliku.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Nieprawidłowy plik Excel.'}, status=400)

        created, skipped = 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0]).strip() if row[0] else ''
            if not name or name == 'None':
                continue
            _, was_created = Department.objects.get_or_create(name=name)
            if was_created:
                created += 1
            else:
                skipped += 1

        return Response({'created': created, 'skipped': skipped})


class PositionImportView(APIView):
    permission_classes = [IsHROrAdmin]

    def post(self, request):
        import openpyxl
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Nie przesłano pliku.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Nieprawidłowy plik Excel.'}, status=400)

        created, skipped = 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0]).strip() if row[0] else ''
            if not name or name == 'None':
                continue
            _, was_created = Position.objects.get_or_create(name=name)
            if was_created:
                created += 1
            else:
                skipped += 1

        return Response({'created': created, 'skipped': skipped})


# ── Companies ─────────────────────────────────────────────────────────────────

class CompanyListCreateView(generics.ListCreateAPIView):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsHROrAdmin()]
        return [IsAuthenticated()]


class CompanyDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    permission_classes = [IsHROrAdmin]


# ── Regions ───────────────────────────────────────────────────────────────────

class RegionListCreateView(generics.ListCreateAPIView):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsHROrAdmin()]
        return [IsAuthenticated()]


class RegionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer
    permission_classes = [IsHROrAdmin]


# ── Contracts ─────────────────────────────────────────────────────────────────

class ContractListCreateView(generics.ListCreateAPIView):
    serializer_class = ContractSerializer
    permission_classes = [IsHROrAdmin]

    def get_queryset(self):
        return Contract.objects.filter(employee_id=self.kwargs['user_pk'])

    def perform_create(self, serializer):
        employee = generics.get_object_or_404(User, pk=self.kwargs['user_pk'])
        serializer.save(employee=employee)


class ContractDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ContractSerializer
    permission_classes = [IsHROrAdmin]

    def get_queryset(self):
        return Contract.objects.filter(employee_id=self.kwargs['user_pk'])


# ── Employee Excel Import ──────────────────────────────────────────────────────

USER_FIELD_MAP = {
    'first_name': 'Imię',
    'last_name': 'Nazwisko',
    'email': 'Email',
    'username': 'Login',
    'position': 'Stanowisko',
    'phone': 'Telefon',
    'hire_date': 'Data zatrudnienia',
    'contract_type': 'Typ umowy',
    'contract_start': 'Umowa od',
    'contract_end': 'Umowa do',
}

class UserImportParseView(APIView):
    """Step 1: upload file, return column names for mapping."""
    permission_classes = [IsHROrAdmin]

    def post(self, request):
        import openpyxl
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Nie przesłano pliku.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Nieprawidłowy plik Excel.'}, status=400)
        headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        preview = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i >= 3:
                break
            preview.append([str(v or '') for v in row])
        return Response({'columns': headers, 'preview': preview, 'field_options': USER_FIELD_MAP})


class UserImportConfirmView(APIView):
    """Step 2: receive mapping, import users."""
    permission_classes = [IsHROrAdmin]

    def post(self, request):
        import openpyxl
        from datetime import datetime
        file = request.FILES.get('file')
        mapping = request.data.get('mapping', {})  # { col_index: field_name }
        default_password = request.data.get('default_password', 'Pracownik1234!')
        default_role = request.data.get('default_role', 'employee')

        if not file:
            return Response({'detail': 'Nie przesłano pliku.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Nieprawidłowy plik Excel.'}, status=400)

        created, skipped, errors = 0, 0, []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {}
            for col_str, field in mapping.items():
                col_idx = int(col_str)
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    data[field] = str(val).strip()

            if not data.get('last_name'):
                continue

            # Auto-generate username if not provided
            if not data.get('username'):
                base = f"{(data.get('first_name', '') or '')[:1].lower()}{(data.get('last_name', '') or '').lower().replace(' ', '')}"
                username = base
                n = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base}{n}"
                    n += 1
                data['username'] = username

            if User.objects.filter(username=data['username']).exists():
                skipped += 1
                continue

            # Parse dates
            for date_field in ('hire_date', 'contract_start', 'contract_end'):
                if data.get(date_field):
                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            data[date_field] = datetime.strptime(data[date_field], fmt).date()
                            break
                        except ValueError:
                            pass
                    else:
                        data.pop(date_field, None)

            try:
                user = User(
                    username=data.get('username', ''),
                    first_name=data.get('first_name', ''),
                    last_name=data.get('last_name', ''),
                    email=data.get('email', ''),
                    position=data.get('position', ''),
                    phone=data.get('phone', ''),
                    role=default_role,
                    hire_date=data.get('hire_date'),
                    contract_type=data.get('contract_type', ''),
                    contract_start=data.get('contract_start'),
                    contract_end=data.get('contract_end'),
                )
                user.set_password(default_password)
                user.save()
                created += 1
            except Exception as e:
                errors.append(f'Wiersz {row_num}: {e}')

        return Response({'created': created, 'skipped': skipped, 'errors': errors})
